import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Acer\\Documents\\exceldemo.xlsx")

sheet = book['Sheet1']
Dict = {}
cell1 = sheet.cell(row=1, column=2).value
cell2 = sheet.cell(row=1, column=3)
cell3 = [sheet.cell(row=2, column=2).value, sheet.cell(row=2, column=3).value]
cell4 = sheet.cell(row=2, column=3).value
sheet.cell(row=9, column=2).value = "rahul"

print(cell1)
print(cell2.value)
print(cell3, cell4)
print(sheet.cell(row=9, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['B5'].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase1":    #to select particular testcase data
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
