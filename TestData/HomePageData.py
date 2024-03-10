import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname": "Rahul", "lastname": "shetty", "gender": "male"}, {"firstname": "ankita", "lastname": "shetty", "gender": "female"}]

    @staticmethod
    def gettestdata(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\Acer\\Documents\\exceldemo.xlsx")
        sheet = book['Sheet1']
        Dict = {}
        for i in range(1, sheet.max_row + 1):   # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:  # to select particular testcase data
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
